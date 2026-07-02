"""
Excel 数据存储 — 从 MediaCrawler ExcelStoreBase 适配

将采集数据导出为格式化 Excel 工作簿，包含：
- 样式化列头（深色背景 + 白色粗体）
- 自动列宽
- 数值格式化
- 多 sheet 分类

使用:
    store = ExcelDataStore("./output/reports")
    result = await collector.collect()
    await store.save(result)

依赖: openpyxl (pip install openpyxl)
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from content_aggregator.sources.collectors.base_collector import SourceResult
from .abstract import DataStore

logger = logging.getLogger(__name__)

# ─── Styles ───────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

COL_WIDTH_MIN = 10
COL_WIDTH_MAX = 60
COL_WIDTH_DEFAULT = 20


class ExcelDataStore(DataStore):
    """Excel 文件导出存储。

    将采集结果导出为格式化 Excel 工作簿，每个 save() 调用创建一个新文件。
    文件命名: {source_name}_{timestamp}.xlsx
    """

    def __init__(self, output_dir: str = "./output/reports", **kwargs):
        if not EXCEL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl"
            )
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _filename(self, source_name: str) -> str:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return self.output_dir / f"{source_name}_{ts}.xlsx"

    def _make_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet
        return wb

    def _write_headers(self, ws: openpyxl.worksheet.Worksheet, headers: List[str]) -> None:
        for col, name in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

    def _write_row(self, ws: openpyxl.worksheet.Worksheet, headers: List[str], data: dict) -> None:
        row_num = ws.max_row + 1
        for col, name in enumerate(headers, 1):
            value = data.get(name, "")
            if isinstance(value, (list, dict)):
                value = str(value)
            elif value is None:
                value = ""
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

    def _auto_width(self, ws: openpyxl.worksheet.Worksheet, headers: List[str]) -> None:
        for col, name in enumerate(headers, 1):
            col_letter = get_column_letter(col)
            # Header width
            w = len(str(name)) + 2
            # Check data cells (first 100 rows max)
            for row in ws.iter_rows(
                min_row=2, max_row=min(ws.max_row, 101),
                min_col=col, max_col=col
            ):
                for cell in row:
                    if cell.value:
                        w = max(w, min(len(str(cell.value)), COL_WIDTH_MAX))
            ws.column_dimensions[col_letter].width = max(min(w, COL_WIDTH_MAX), COL_WIDTH_MIN)

    async def save(self, result: SourceResult) -> str:
        """将 SourceResult 导出为 Excel 文件。

        Returns:
            Excel 文件路径
        """
        articles = result.data
        if not articles:
            logger.info("[ExcelDataStore] 无数据，跳过导出")
            return ""

        wb = self._make_workbook()
        headers = self._infer_headers(articles)

        ws = wb.create_sheet(title="Articles")
        self._write_headers(ws, headers)
        for article in articles:
            self._write_row(ws, headers, article)

        self._auto_width(ws, headers)

        filepath = self._filename(result.source_name)
        wb.save(str(filepath))
        logger.info("[ExcelDataStore] 已导出 %d 条到 %s", len(articles), filepath)
        return str(filepath)

    def _infer_headers(self, articles: List[dict]) -> List[str]:
        """从文章数据推断列顺序（统一排序，缺失列自动补充）"""
        # 定义标准列名（优先显示）
        standard = [
            "title", "content", "author", "source", "published_at",
            "summary", "url", "tags",
        ]
        # 收集所有实际存在的 key
        seen = set()
        for a in articles:
            seen.update(a.keys())
        # 按标准顺序 + 剩余按字母序
        ordered = [h for h in standard if h in seen]
        remaining = sorted(h for h in seen if h not in standard)
        return ordered + remaining

    # ── DataStore 接口兼容 ──────────────────────────────────────────
    # Excel 格式不适合作为查询后端，以下方法抛出 NotImplementedError
    # 说明用户应将数据导入 JSON/DB store 后再查询。

    async def query(self, source_name: Optional[str] = None,
                    limit: int = 100, offset: int = 0) -> List[Dict[str, Any]]:
        raise NotImplementedError(
            "ExcelDataStore 不支持查询。使用 JsonDataStore 或 DbDataStore "
            "进行数据检索，Excel 仅用于导出。"
        )

    async def count(self, source_name: Optional[str] = None) -> int:
        raise NotImplementedError(
            "ExcelDataStore 不支持计数。使用 JsonDataStore 或 DbDataStore。"
        )
